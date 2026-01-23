#!/usr/bin/env python3
"""
MNTD Reagent Inventory Excel to SQL Converter

Reads the MNTD Reagent inventory Excel file, normalizes data, and generates
production-safe SQL scripts for inventory import without using Liquibase.

Usage:
    python3 excel_to_sql.py --input "MNTD Reagent inventory (2) (1).xlsx" --output ./sql
"""

import sys
import json
import uuid
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.utils import get_column_letter


class InventoryImporter:
    """Handles Excel parsing and SQL generation for inventory import."""

    # Valid UOM values in system
    VALID_UOMS = {
        'vial', 'ml', 'ul', 'mg', 'ug', 'nmol', 'pack', 'box', 'lyophilized',
        'unit', 'ug/ml', 'u/ml', 'percent', 'mm', 'cm', 'g', 'ng', 'pmol'
    }

    # Sheet to item type mapping
    SHEET_TYPE_MAP = {
        'Enzymes': 'ENZYME',
        'Immunoassay': 'REAGENT',
        'Digital PCR': 'REAGENT',
        'qPCR': 'REAGENT',
    }

    # Common column headers to standardize
    COLUMN_MAPPING = {
        'SN': 'serial_number',
        'S.N': 'serial_number',
        '   SN': 'serial_number',
        'Ref No.': 'ref_number',
        'Ref/ LOT No.': 'ref_number',
        'Ref/LOT No.': 'ref_number',
        'Ref No': 'ref_number',
        'Vial No.': 'vial_number',
        'Items': 'item_name',
        'Experiment type': 'experiment_type',
        'Concetration': 'concentration',
        'Concentration': 'concentration',
        'concentration': 'concentration',
        'Quantity': 'quantity',
        'Unit': 'unit',
        'Volume': 'volume',
        'volume': 'volume',
        'unit': 'volume_unit',
        'Manufacturing date': 'manufacturing_date',
        'Open date': 'open_date',
        'Expiry date': 'expiry_date',
        'Box number': 'box_number',
        'Box No.': 'box_number',
        'BOX Number': 'box_number',
        'box number': 'box_number',
        'BOX NUMBER': 'box_number',
        'REMARK': 'remark',
        'Remark': 'remark',
        'remark': 'remark',
        'Remarks': 'remark',
        'Recived from': 'received_from',
        'Received from': 'received_from',
        'Label and location': 'label_location',
        'Storage condition': 'storage_condition',
        'Requested by': 'requested_by',
        'Refridgrator #': 'refrigerator',
        'Freezer partition': 'freezer_partition',
        'storage temprature & condition': 'storage_condition',
    }

    def __init__(self, excel_file: str, output_dir: str):
        self.excel_file = Path(excel_file)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.items_inserted = {}  # {item_name: item_id} cache
        self.items_sql = []  # SQL INSERT statements for items
        self.lots_sql = []  # SQL INSERT statements for lots
        self.transactions_sql = []  # SQL for audit trail

        self.stats = {
            'total_rows_read': 0,
            'items_created': 0,
            'lots_created': 0,
            'rows_skipped': 0,
            'errors': [],
            'warnings': [],
        }

    def normalize_uom(self, uom: Optional[str]) -> str:
        """Normalize Unit of Measure variations."""
        if not uom:
            return 'unit'

        uom = str(uom).strip().lower()

        # Exact matches
        if uom in self.VALID_UOMS:
            return uom

        # Normalization rules
        uom_map = {
            'milliliter': 'ml',
            'microliter': 'ul',
            'μl': 'ul',
            'µl': 'ul',
            'microgram': 'ug',
            'μg': 'ug',
            'µg': 'ug',
            'milligram': 'mg',
            'nanomol': 'nmol',
            'picomol': 'pmol',
            'stock': 'unit',
            'vials': 'vial',
            'packs': 'pack',
            'boxes': 'box',
            'lyophilized vial': 'lyophilized',
            'nmol': 'nmol',
            'u/ml': 'u/ml',
            'ug/ml': 'ug/ml',
        }

        if uom in uom_map:
            return uom_map[uom]

        # If still not found, return original or default
        return 'unit'

    def normalize_date(self, date_val) -> Optional[str]:
        """Convert date to PostgreSQL format (YYYY-MM-DD)."""
        if not date_val:
            return None

        if hasattr(date_val, 'strftime'):
            # datetime object
            return date_val.strftime('%Y-%m-%d')

        # Try string parsing
        date_str = str(date_val).strip()
        if not date_str or date_str.lower() in ['na', 'n/a', 'none']:
            return None

        # Try common date formats
        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y']:
            try:
                parsed = datetime.strptime(date_str, fmt)
                return parsed.strftime('%Y-%m-%d')
            except ValueError:
                continue

        # Could not parse
        return None

    def normalize_quantity(self, qty: Optional) -> float:
        """Convert quantity to float."""
        if qty is None:
            return 1.0

        try:
            q = str(qty).strip()
            # Handle ranges like "23, 33, 16"
            if ',' in q:
                parts = [float(p.strip()) for p in q.split(',')]
                return sum(parts)
            return float(q)
        except (ValueError, AttributeError):
            return 1.0

    def parse_row(self, row: List, headers: List[str], sheet_name: str) -> Optional[Dict]:
        """Parse a single Excel row into normalized dictionary."""
        record = {}

        for col_idx, cell_val in enumerate(row):
            if col_idx >= len(headers):
                break

            header = headers[col_idx]
            if header not in self.COLUMN_MAPPING:
                continue

            key = self.COLUMN_MAPPING[header]
            record[key] = cell_val

        # Skip if no item name
        if not record.get('item_name'):
            self.stats['rows_skipped'] += 1
            return None

        # Required fields with defaults
        record['sheet_name'] = sheet_name
        record['item_type'] = self.SHEET_TYPE_MAP.get(sheet_name, 'REAGENT')

        return record

    def generate_item_sql(self, item_name: str, item_type: str, record: Dict) -> Tuple[str, str]:
        """Generate SQL INSERT for inventory_item."""
        fhir_uuid = str(uuid.uuid4())
        category = record.get('experiment_type') or 'General'
        project_name = 'MNTD Research'

        # Extract units from various fields
        units = self.normalize_uom(record.get('volume_unit') or record.get('unit'))

        # Escape single quotes in text
        item_name_escaped = item_name.replace("'", "''")
        category_escaped = str(category).replace("'", "''")

        sql = f"""INSERT INTO inventory_item (
            fhir_uuid, name, item_type, category, units, is_active,
            project_name, created_date, low_stock_threshold, expiration_alert_days, version
        ) VALUES (
            '{fhir_uuid}', '{item_name_escaped}', '{item_type}', '{category_escaped}',
            '{units}', 'Y', '{project_name}', NOW(), 100, 30, 1
        ) ON CONFLICT (fhir_uuid) DO NOTHING
        RETURNING id;"""

        return sql, fhir_uuid

    def generate_lot_sql(self, item_id: str, record: Dict, fhir_uuid: str) -> Tuple[str, str]:
        """Generate SQL INSERT for inventory_lot."""
        lot_fhir_uuid = str(uuid.uuid4())

        # Lot number is critical
        lot_number = record.get('ref_number') or f"LOT_{datetime.now().strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"
        lot_number = str(lot_number).replace("'", "''")

        # Dates
        manufacturing_date = self.normalize_date(record.get('manufacturing_date'))
        open_date = self.normalize_date(record.get('open_date'))
        expiry_date = self.normalize_date(record.get('expiry_date'))

        # If no expiry date, calculate 24 months from manufacturing or today
        if not expiry_date:
            if manufacturing_date:
                expiry = datetime.strptime(manufacturing_date, '%Y-%m-%d') + timedelta(days=730)
            else:
                expiry = datetime.now() + timedelta(days=730)
            expiry_date = expiry.strftime('%Y-%m-%d')

        # Quantities
        quantity = self.normalize_quantity(record.get('quantity'))

        # Status logic
        status = 'ACTIVE'
        if expiry_date:
            expiry_dt = datetime.strptime(expiry_date, '%Y-%m-%d')
            if expiry_dt < datetime.now():
                status = 'EXPIRED'
        elif str(record.get('expiry_date', '')).lower() in ['na', 'n/a']:
            status = 'ACTIVE'

        qc_status = 'PENDING'

        # Unit size / description
        unit_size = record.get('volume') or ''
        unit_size = str(unit_size).replace("'", "''")

        # Remarks - combine all relevant fields
        remarks_parts = []
        if record.get('concentration'):
            remarks_parts.append(f"Concentration: {record.get('concentration')}")
        if record.get('vial_number'):
            remarks_parts.append(f"Vial #: {record.get('vial_number')}")
        if record.get('box_number'):
            remarks_parts.append(f"Box: {record.get('box_number')}")
        if record.get('storage_condition'):
            remarks_parts.append(f"Storage: {record.get('storage_condition')}")
        if record.get('remark'):
            remarks_parts.append(f"Note: {record.get('remark')}")

        remarks = ' | '.join(remarks_parts)
        remarks = remarks[:1000].replace("'", "''")  # Max 1000 chars

        # Location - try to use box_number or storage location
        storage_path = record.get('box_number') or record.get('label_location') or 'To Be Determined'
        storage_path = str(storage_path).replace("'", "''")[:255]

        sql = f"""INSERT INTO inventory_lot (
            fhir_uuid, inventory_item_id, lot_number, initial_quantity, current_quantity,
            expiration_date, date_opened, qc_status, status, unit_size, storage_path,
            remarks, created_date, version
        ) VALUES (
            '{lot_fhir_uuid}', {item_id}, '{lot_number}', {quantity}, {quantity},
            '{expiry_date}', {f"'{open_date}'" if open_date else 'NULL'},
            '{qc_status}', '{status}', '{unit_size}', '{storage_path}',
            '{remarks}', NOW(), 1
        ) RETURNING id;"""

        return sql, lot_fhir_uuid

    def generate_transaction_sql(self, lot_id: str, quantity: float) -> str:
        """Generate SQL INSERT for inventory_transaction (audit trail)."""
        sql = f"""INSERT INTO inventory_transaction (
            lot_id, transaction_type, quantity_change, quantity_after,
            transaction_date, notes, created_date, version
        ) VALUES (
            {lot_id}, 'RECEIPT', {quantity}, {quantity},
            NOW(), 'Batch import from MNTD Excel file', NOW(), 1
        );"""
        return sql

    def process_sheet(self, ws, sheet_name: str) -> None:
        """Process a single Excel sheet."""
        print(f"  Processing sheet: {sheet_name}")

        # Get headers (first row with values)
        headers = []
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            if row[0]:  # First column has value
                headers = [h for h in row]
                header_row = row_idx
                break

        if not headers:
            print(f"    ⚠ No headers found, skipping")
            return

        print(f"    Headers found at row {header_row}: {len([h for h in headers if h])} columns")

        # Process data rows
        data_start = header_row + 1
        row_count = 0

        for row_idx in range(data_start, ws.max_row + 1):
            row = []
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row.append(cell.value)

            # Skip empty rows
            if not any(row):
                continue

            self.stats['total_rows_read'] += 1
            row_count += 1

            # Parse row
            record = self.parse_row(row, headers, sheet_name)
            if not record:
                continue

            item_name = str(record['item_name']).strip()
            item_type = record['item_type']

            # Generate item SQL if not exists
            if item_name not in self.items_inserted:
                item_sql, item_uuid = self.generate_item_sql(item_name, item_type, record)
                self.items_sql.append(item_sql)
                # Use placeholder for item_id - will need CTE or wrapping
                item_id = f"(SELECT id FROM inventory_item WHERE fhir_uuid = '{item_uuid}')"
                self.items_inserted[item_name] = item_uuid
                self.stats['items_created'] += 1
            else:
                item_uuid = self.items_inserted[item_name]
                item_id = f"(SELECT id FROM inventory_item WHERE fhir_uuid = '{item_uuid}')"

            # Generate lot SQL
            lot_sql, lot_uuid = self.generate_lot_sql(item_id, record, item_uuid)

            # Wrap lot insert to capture ID for transaction
            lot_id = f"(SELECT id FROM inventory_lot WHERE fhir_uuid = '{lot_uuid}')"

            self.lots_sql.append(lot_sql)

            # Transaction SQL
            qty = self.normalize_quantity(record.get('quantity'))
            trans_sql = self.generate_transaction_sql(lot_id, qty)
            self.transactions_sql.append(trans_sql)

            self.stats['lots_created'] += 1

            if row_count % 100 == 0:
                print(f"    Processed {row_count} rows...")

        print(f"    ✓ Completed {row_count} rows")

    def process_excel(self) -> None:
        """Process all sheets in Excel file."""
        print(f"Reading Excel file: {self.excel_file}")

        wb = openpyxl.load_workbook(self.excel_file, data_only=False)
        sheet_count = len(wb.sheetnames)
        print(f"Found {sheet_count} sheets\n")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.process_sheet(ws, sheet_name)

        print(f"\n✓ Excel processing complete")
        print(f"  Total rows read: {self.stats['total_rows_read']}")
        print(f"  Unique items: {self.stats['items_created']}")
        print(f"  Lots created: {self.stats['lots_created']}")

    def write_sql_files(self) -> None:
        """Write generated SQL to files."""
        print(f"\nGenerating SQL files...")

        # Items SQL file
        items_file = self.output_dir / '02-insert-inventory-items.sql'
        with open(items_file, 'w') as f:
            f.write("-- MNTD Inventory Import: Inventory Items\n")
            f.write(f"-- Generated: {datetime.now().isoformat()}\n")
            f.write("-- Total items: {}\n\n".format(self.stats['items_created']))
            f.write("BEGIN TRANSACTION;\n\n")
            for sql in self.items_sql:
                f.write(sql + "\n")
            f.write("\nCOMMIT;\n")
        print(f"  ✓ {items_file} ({len(self.items_sql)} inserts)")

        # Lots SQL file
        lots_file = self.output_dir / '03-insert-inventory-lots.sql'
        with open(lots_file, 'w') as f:
            f.write("-- MNTD Inventory Import: Inventory Lots & Transactions\n")
            f.write(f"-- Generated: {datetime.now().isoformat()}\n")
            f.write("-- Total lots: {}\n\n".format(self.stats['lots_created']))
            f.write("BEGIN TRANSACTION;\n\n")

            # Include both lots and transactions together
            for lot_sql, trans_sql in zip(self.lots_sql, self.transactions_sql):
                f.write(lot_sql + "\n")
                f.write(trans_sql + "\n")

            f.write("\nCOMMIT;\n")
        print(f"  ✓ {lots_file} ({len(self.lots_sql)} inserts + {len(self.transactions_sql)} transactions)")

    def write_report(self) -> None:
        """Write summary report."""
        report_file = self.output_dir / 'import-report.json'

        report = {
            'generated_at': datetime.now().isoformat(),
            'excel_file': str(self.excel_file),
            'statistics': self.stats,
            'output_files': {
                'items_sql': str(self.output_dir / '02-insert-inventory-items.sql'),
                'lots_sql': str(self.output_dir / '03-insert-inventory-lots.sql'),
            }
        }

        with open(report_file, 'w') as f:
            json.dump(report, f, indent=2)

        print(f"  ✓ {report_file}")

    def run(self) -> bool:
        """Run the complete import process."""
        print("=" * 70)
        print("MNTD Inventory Excel to SQL Converter")
        print("=" * 70)

        try:
            self.process_excel()
            self.write_sql_files()
            self.write_report()

            print("\n" + "=" * 70)
            print("✓ SUCCESS - SQL generation complete")
            print("=" * 70)
            print(f"\nNext steps:")
            print(f"  1. Review SQL files in: {self.output_dir}")
            print(f"  2. Backup your database")
            print(f"  3. Run: psql -U postgres -d openelis < {self.output_dir}/02-insert-inventory-items.sql")
            print(f"  4. Run: psql -U postgres -d openelis < {self.output_dir}/03-insert-inventory-lots.sql")
            print(f"  5. Verify in OpenELIS UI: /inventory → Catalog\n")

            return True
        except Exception as e:
            print(f"\n✗ ERROR: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    parser = argparse.ArgumentParser(
        description='Convert MNTD Reagent Inventory Excel to SQL scripts'
    )
    parser.add_argument('--input', required=True, help='Path to Excel file')
    parser.add_argument('--output', required=True, help='Output directory for SQL files')

    args = parser.parse_args()

    # Verify Excel file exists
    if not Path(args.input).exists():
        print(f"✗ Error: Excel file not found: {args.input}")
        sys.exit(1)

    importer = InventoryImporter(args.input, args.output)
    success = importer.run()
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
